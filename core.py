"""
core.py — Moteur comptable (sans interface graphique).

Toute la logique métier vit ici, indépendamment de Tkinter, pour rester
testable en ligne de commande. main.py ne fait qu'appeler ces fonctions.
"""
import json
import os
import sys
import sqlite3
from datetime import date


def _resource_dir():
    """Dossier des ressources bundlées : gère le cas exécutable PyInstaller."""
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# Comptes SYSCOHADA "spéciaux" utilisés par les calculs automatiques
# (repris de la maquette Excel d'origine).
# ---------------------------------------------------------------------------
COMPTES_STOCK = ["310000", "320000", "331000", "360000"]
COMPTES_TRESORERIE = ["521000", "531000", "570000", "585000"]
COMPTES_CAPITAL = ["101000", "118000", "121000"]
COMPTE_SUBVENTIONS = "141000"
COMPTE_PROVISIONS = "191000"
COMPTES_DETTES_FIN = ["162000", "165000"]
COMPTES_PRODUITS_EXPL = ["701000", "702000", "705000", "706000"]
COMPTE_SUBV_EXPL = "710000"
COMPTE_AUTRES_PRODUITS = "758000"
COMPTES_ACHATS = ["601000", "602000", "604000", "605000"]
COMPTES_TRANSPORT = ["610000", "614000"]
COMPTES_SERVICES_EXT = ["622000", "624000", "625000", "626000", "627000", "628000",
                         "631000", "632000", "633000"]
COMPTES_IMPOTS = ["641000", "645000"]
COMPTE_AUTRES_CHARGES = "651000"
COMPTES_PERSONNEL = ["661000", "663000", "664000"]
COMPTES_DOTATIONS = ["681000", "691000"]
COMPTES_PRODUITS_FIN = ["771000", "776000"]
COMPTES_CHARGES_FIN = ["671000", "676000"]

# ---------------------------------------------------------------------------
# Liasse fiscale — codes SYSCOHADA "système normal" (BILAN / RESULTAT)
# NB : les totaux (AD, AI, AZ, BK, BT, BZ, CP, DD, DP, DT, DZ) sont fiables
# (dérivés directement de la partie double). Le détail par ligne (AE..AN,
# CA/CH/CJ, DA/DJ/DK/DM/DR) est une répartition indicative par plage de
# comptes — à vérifier avec votre expert-comptable avant tout dépôt officiel.
# ---------------------------------------------------------------------------
RANGES_INCORP = {"AE": (211000, 211999), "AF": (212000, 214999),
                  "AG": (215000, 216999), "AH": (217000, 219999)}
RANGE_AMORT_INCORP = (281000, 281999)
RANGES_CORP = {"AJ": [(220000, 229999)], "AK": [(230000, 233999)],
               "AL": [(234000, 239999)],
               "AM": [(240000, 244999), (246000, 249999)],
               "AN": [(245000, 245999)]}
RANGE_AMORT_CORP = (282000, 297999)
RANGE_AVANCES_IMMO = (250000, 252999)
RANGE_TITRES_PARTICIPATION = (260000, 268999)
RANGE_AUTRES_IMMO_FIN = (270000, 278999)

RANGES_CAPITAUX = {"CA": [(101000, 104999)], "CD": [(105000, 105999)],
                    "CF_CG": [(110000, 118999)], "CH": [(120000, 129999)],
                    "CL": [(140000, 148999)], "CM": [(150000, 158999)]}
RANGE_DETTES_FIN = (160000, 168999)
RANGE_DETTES_LOCATION = (170000, 178999)
RANGE_PROVISIONS_RC = (190000, 198999)

RANGE_STOCKS = (300000, 399999)
RANGE_AVANCES_FOURN = (409000, 409999)
RANGE_CLIENTS = (411000, 419999)
RANGE_FOURNISSEURS = (401000, 408999)
RANGE_DETTES_FISC_SOC = (420000, 449999)
RANGE_AUTRES_DETTES = (450000, 499999)


def _in_ranges(code_int, ranges):
    if isinstance(ranges, tuple):
        ranges = [ranges]
    return any(lo <= code_int <= hi for lo, hi in ranges)


def _sum_range(balance, ranges, classe=None):
    total = 0.0
    for b in balance:
        code_int = int(b["code"])
        if classe and b["classe"] != classe:
            continue
        if _in_ranges(code_int, ranges):
            total += b["solde_cloture"]
    return total


def compute_liasse_bilan(conn, stock_initial=0.0):
    """Bilan au format SYSCOHADA système normal (codes officiels)."""
    balance = compute_balance(conn, only_with_movement=False)
    bilan_simple = compute_bilan(conn, stock_initial=stock_initial)

    # --- Détail indicatif Immobilisations incorporelles ---
    incorp_brut = {k: _sum_range(balance, [rng]) for k, rng in RANGES_INCORP.items()}
    total_incorp_brut = sum(incorp_brut.values())
    amort_incorp_total = -_sum_range(balance, [RANGE_AMORT_INCORP])  # positif
    incorp_net = {}
    for k, brut in incorp_brut.items():
        part = (brut / total_incorp_brut * amort_incorp_total) if total_incorp_brut else 0
        incorp_net[k] = brut - part

    # --- Détail indicatif Immobilisations corporelles ---
    corp_brut = {k: _sum_range(balance, rngs) for k, rngs in RANGES_CORP.items()}
    total_corp_brut = sum(corp_brut.values())
    amort_corp_total = -_sum_range(balance, [RANGE_AMORT_CORP])
    corp_net = {}
    for k, brut in corp_brut.items():
        part = (brut / total_corp_brut * amort_corp_total) if total_corp_brut else 0
        corp_net[k] = brut - part

    avances_immo = _sum_range(balance, [RANGE_AVANCES_IMMO])
    titres_participation = _sum_range(balance, [RANGE_TITRES_PARTICIPATION])
    autres_immo_fin = _sum_range(balance, [RANGE_AUTRES_IMMO_FIN])

    # --- Détail indicatif Capitaux propres ---
    capitaux_detail = {k: -_sum_range(balance, rngs) for k, rngs in RANGES_CAPITAUX.items()}
    dettes_financieres = -_sum_range(balance, [RANGE_DETTES_FIN])
    dettes_location = -_sum_range(balance, [RANGE_DETTES_LOCATION])
    provisions_rc = -_sum_range(balance, [RANGE_PROVISIONS_RC])

    # --- Détail indicatif Passif circulant ---
    fournisseurs = -_sum_range(balance, [RANGE_FOURNISSEURS])
    avances_fourn = -_sum_range(balance, [RANGE_AVANCES_FOURN])
    dettes_fisc_soc = -_sum_range(balance, [RANGE_DETTES_FISC_SOC])
    autres_dettes = -_sum_range(balance, [RANGE_AUTRES_DETTES])

    # --- Détail indicatif Actif circulant ---
    avances_versees = _sum_range(balance, [RANGE_AVANCES_FOURN])
    clients = _sum_range(balance, [RANGE_CLIENTS])

    return {
        "totaux": bilan_simple,
        "actif_detail": {
            **{k: {"brut": incorp_brut[k], "net": incorp_net[k]} for k in incorp_brut},
            **{k: {"brut": corp_brut[k], "net": corp_net[k]} for k in corp_brut},
            "AP": {"brut": avances_immo, "net": avances_immo},
            "AR": {"brut": titres_participation, "net": titres_participation},
            "AS": {"brut": autres_immo_fin, "net": autres_immo_fin},
        },
        "actif_circulant_detail": {
            "BH": avances_versees, "BI": clients,
        },
        "passif_detail": {
            **capitaux_detail,
            "DA": dettes_financieres, "DB": dettes_location, "DC": provisions_rc,
            "DJ": fournisseurs, "DH_avances": avances_fourn,
            "DK": dettes_fisc_soc, "DM": autres_dettes,
        },
    }


def compute_liasse_resultat(conn):
    """Compte de résultat au format SYSCOHADA système normal (codes officiels)."""
    balance = compute_balance(conn, only_with_movement=False)

    def net_produit(codes):
        d, c = _sum_accounts(balance, codes)
        return c - d

    def net_charge(codes):
        d, c = _sum_accounts(balance, codes)
        return d - c

    ta = net_produit(["701000"])
    ra = net_charge(["601000"])
    xa = ta - ra  # marge commerciale (simplifiée, sans variation de stock marchandises isolée)

    tb = net_produit(["702000"])
    tc = net_produit(["705000", "706000"])
    td = 0.0
    xb = ta + tb + tc + td

    stock_d, stock_c = _sum_accounts(balance, ["360000"])
    te = stock_d - stock_c
    th = net_produit(["758000"])
    tg = net_produit(["710000"])

    rc = net_charge(["602000"])
    re = net_charge(["604000", "605000"])
    rg = net_charge(["610000", "614000"])
    rh = net_charge(["622000", "624000", "625000", "626000", "627000", "628000",
                      "631000", "632000", "633000"])
    ri = net_charge(["641000", "645000"])
    rj = net_charge(["651000"])
    xc = xb + (-ra) + te + tg + th + (-rc) + (-re) + (-rg) + (-rh) + (-ri) + (-rj)

    rk = net_charge(["661000", "663000", "664000"])
    xd = xc - rk

    rl = net_charge(["681000", "691000"])
    xe = xd - rl

    tk = net_produit(["771000", "776000"])
    rm = net_charge(["671000", "676000"])
    xf = tk - rm
    xg = xe + xf

    xh = 0.0  # Résultat HAO — non tracé dans cette application
    rq = 0.0  # Participation des travailleurs — non tracée
    rs = 0.0  # Impôt sur le résultat — non tracé (IS à calculer/saisir séparément)
    xi = xg + xh + rq + rs

    return {
        "TA": ta, "RA": ra, "XA": xa,
        "TB": tb, "TC": tc, "TD": td, "XB": xb,
        "TE": te, "TG": tg, "TH": th,
        "RC": rc, "RE": re, "RG": rg, "RH": rh, "RI": ri, "RJ": rj, "XC": xc,
        "RK": rk, "XD": xd,
        "RL": rl, "XE": xe,
        "TK": tk, "RM": rm, "XF": xf, "XG": xg,
        "XH": xh, "RQ": rq, "RS": rs, "XI": xi,
    }


def default_db_path():
    """Emplacement du fichier de données, à côté de l'exécutable."""
    base = os.environ.get("LOCALAPPDATA") or os.path.expanduser("~")
    folder = os.path.join(base, "SaisieComptable")
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, "comptabilite.db")


def get_connection(db_path=None):
    db_path = db_path or default_db_path()
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    init_db(conn)
    return conn


def init_db(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS accounts (
            code TEXT PRIMARY KEY,
            label TEXT NOT NULL,
            classe TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            piece TEXT,
            journal TEXT,
            compte TEXT NOT NULL,
            tiers TEXT,
            libelle TEXT,
            debit REAL NOT NULL DEFAULT 0,
            credit REAL NOT NULL DEFAULT 0,
            flux_code TEXT,
            analytic_code TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS opening_balances (
            code TEXT PRIMARY KEY,
            solde REAL NOT NULL DEFAULT 0
        )
    """)
    conn.commit()
    _migrate(conn)
    if conn.execute("SELECT COUNT(*) FROM accounts").fetchone()[0] == 0:
        load_plan_comptable(conn)


def _migrate(conn):
    """Ajoute les colonnes/tables manquantes si la base a été créée par une version antérieure."""
    cols = [r["name"] for r in conn.execute("PRAGMA table_info(entries)")]
    if "analytic_code" not in cols:
        conn.execute("ALTER TABLE entries ADD COLUMN analytic_code TEXT")
    # Migre l'ancien mécanisme "stock_initial_<compte>" (settings) vers opening_balances
    old_rows = conn.execute("SELECT key, value FROM settings WHERE key LIKE 'stock_initial_%'").fetchall()
    for row in old_rows:
        code = row["key"].replace("stock_initial_", "")
        try:
            val = float(row["value"])
        except (TypeError, ValueError):
            val = 0.0
        if val:
            conn.execute("INSERT OR REPLACE INTO opening_balances (code, solde) VALUES (?, ?)", (code, val))
        conn.execute("DELETE FROM settings WHERE key = ?", (row["key"],))
    conn.commit()


def get_setting(conn, key, default=0.0):
    row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    return float(row["value"]) if row else default


def set_setting(conn, key, value):
    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, str(value)))
    conn.commit()


# ---------------------------------------------------------------------------
# Soldes d'ouverture (report à nouveau) — un solde signé par compte, saisi
# une fois en début d'exercice. Balance de clôture = solde d'ouverture +
# mouvements de l'exercice (Débit - Crédit).
# ---------------------------------------------------------------------------
def get_opening_balance(conn, code):
    row = conn.execute("SELECT solde FROM opening_balances WHERE code = ?", (code,)).fetchone()
    return row["solde"] if row else 0.0


def set_opening_balance(conn, code, value):
    conn.execute("INSERT OR REPLACE INTO opening_balances (code, solde) VALUES (?, ?)", (code, value))
    conn.commit()


def list_opening_balances(conn):
    rows = conn.execute("""
        SELECT o.code, a.label, a.classe, o.solde
        FROM opening_balances o JOIN accounts a ON a.code = o.code
        WHERE o.solde != 0
        ORDER BY o.code
    """).fetchall()
    return [dict(r) for r in rows]


def total_opening_balance(conn):
    row = conn.execute("SELECT COALESCE(SUM(solde), 0) t FROM opening_balances").fetchone()
    return row["t"]


def load_plan_comptable(conn, json_path=None):
    """Charge le plan comptable (bundlé avec l'application) dans la base."""
    if json_path is None:
        json_path = os.path.join(_resource_dir(), "plan_comptable.json")
    with open(json_path, encoding="utf-8") as f:
        accounts = json.load(f)
    conn.executemany(
        "INSERT OR REPLACE INTO accounts (code, label, classe) VALUES (?, ?, ?)",
        [(a["code"], a["label"], a["classe"]) for a in accounts],
    )
    conn.commit()


# ---------------------------------------------------------------------------
# Comptes
# ---------------------------------------------------------------------------
def search_accounts(conn, query, limit=50):
    query = (query or "").strip()
    if not query:
        rows = conn.execute("SELECT code, label, classe FROM accounts ORDER BY code LIMIT ?", (limit,))
    else:
        like = f"%{query}%"
        rows = conn.execute(
            "SELECT code, label, classe FROM accounts "
            "WHERE code LIKE ? OR label LIKE ? ORDER BY code LIMIT ?",
            (f"{query}%", like, limit),
        )
    return [dict(r) for r in rows]


def get_account_label(conn, code):
    row = conn.execute("SELECT label FROM accounts WHERE code = ?", (code,)).fetchone()
    return row["label"] if row else "Compte introuvable"


# ---------------------------------------------------------------------------
# Écritures (Saisie)
# ---------------------------------------------------------------------------
def add_entry(conn, date_str, piece, journal, compte, tiers, libelle, debit, credit,
              flux_code="", analytic_code=""):
    conn.execute(
        """INSERT INTO entries (date, piece, journal, compte, tiers, libelle, debit, credit,
                                 flux_code, analytic_code)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (date_str, piece, journal, compte, tiers, libelle, debit or 0, credit or 0,
         flux_code, analytic_code),
    )
    conn.commit()


def update_entry(conn, entry_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE entries SET {cols} WHERE id = ?", (*fields.values(), entry_id))
    conn.commit()


def delete_entry(conn, entry_id):
    conn.execute("DELETE FROM entries WHERE id = ?", (entry_id,))
    conn.commit()


def list_entries(conn, order_by="date"):
    rows = conn.execute(f"SELECT * FROM entries ORDER BY {order_by}, id").fetchall()
    return [dict(r) for r in rows]


def totals_debit_credit(conn):
    row = conn.execute("SELECT COALESCE(SUM(debit),0) d, COALESCE(SUM(credit),0) c FROM entries").fetchone()
    return row["d"], row["c"]


# ---------------------------------------------------------------------------
# Balance
# ---------------------------------------------------------------------------
def compute_balance(conn, only_with_movement=True, include_zero_opening=True):
    rows = conn.execute("""
        SELECT a.code, a.label, a.classe,
               COALESCE(SUM(e.debit), 0)  AS debit,
               COALESCE(SUM(e.credit), 0) AS credit
        FROM accounts a
        LEFT JOIN entries e ON e.compte = a.code
        GROUP BY a.code, a.label, a.classe
        ORDER BY a.code
    """).fetchall()
    openings = {r["code"]: r["solde"] for r in conn.execute("SELECT code, solde FROM opening_balances")}
    result = []
    for r in rows:
        debit, credit = r["debit"], r["credit"]
        ouverture = openings.get(r["code"], 0.0)
        if only_with_movement and debit == 0 and credit == 0 and ouverture == 0:
            continue
        solde_mouvement = debit - credit
        result.append({
            "code": r["code"], "label": r["label"], "classe": r["classe"],
            "debit": debit, "credit": credit, "solde": solde_mouvement,
            "solde_ouverture": ouverture,
            "solde_cloture": ouverture + solde_mouvement,
        })
    return result


def _sum_accounts(balance, codes):
    by_code = {b["code"]: b for b in balance}
    debit = sum(by_code[c]["debit"] for c in codes if c in by_code)
    credit = sum(by_code[c]["credit"] for c in codes if c in by_code)
    return debit, credit


def _sum_accounts_cloture(balance, codes):
    """Somme des soldes de CLÔTURE (ouverture + mouvements) pour une liste de comptes."""
    by_code = {b["code"]: b for b in balance}
    return sum(by_code[c]["solde_cloture"] for c in codes if c in by_code)


def _sum_class(balance, classe, sign=None, field="solde_cloture"):
    total = 0
    for b in balance:
        if b["classe"] != classe:
            continue
        v = b[field]
        if sign == "pos" and v <= 0:
            continue
        if sign == "neg" and v >= 0:
            continue
        total += v
    return total


# ---------------------------------------------------------------------------
# Compte de résultat
# ---------------------------------------------------------------------------
def compute_compte_resultat(conn):
    balance = compute_balance(conn, only_with_movement=False)

    def net_produit(codes):
        d, c = _sum_accounts(balance, codes)
        return c - d

    def net_charge(codes):
        d, c = _sum_accounts(balance, codes)
        return d - c

    produits = {
        "Ventes (marchandises, produits finis, travaux, services)": net_produit(COMPTES_PRODUITS_EXPL),
        "Subventions d'exploitation": net_produit([COMPTE_SUBV_EXPL]),
        "Autres produits": net_produit([COMPTE_AUTRES_PRODUITS]),
    }
    total_produits = sum(produits.values())

    charges = {
        "Achats (marchandises et matières)": net_charge(COMPTES_ACHATS),
        "Transports": net_charge(COMPTES_TRANSPORT),
        "Services extérieurs": net_charge(COMPTES_SERVICES_EXT),
        "Impôts et taxes": net_charge(COMPTES_IMPOTS),
        "Autres charges": net_charge([COMPTE_AUTRES_CHARGES]),
        "Charges de personnel": net_charge(COMPTES_PERSONNEL),
        "Dotations aux amortissements et provisions": net_charge(COMPTES_DOTATIONS),
    }
    total_charges = sum(charges.values())

    resultat_exploitation = total_produits - total_charges

    produits_fin = net_produit(COMPTES_PRODUITS_FIN)
    charges_fin = net_charge(COMPTES_CHARGES_FIN)
    resultat_financier = produits_fin - charges_fin

    resultat_net = resultat_exploitation + resultat_financier

    return {
        "produits": produits, "total_produits": total_produits,
        "charges": charges, "total_charges": total_charges,
        "resultat_exploitation": resultat_exploitation,
        "produits_financiers": produits_fin, "charges_financieres": charges_fin,
        "resultat_financier": resultat_financier,
        "resultat_net": resultat_net,
    }


# ---------------------------------------------------------------------------
# Bilan
# ---------------------------------------------------------------------------
def compute_bilan(conn, stock_initial=0.0):
    """stock_initial : conservé pour compatibilité, normalement inutile désormais —
    utilisez la table des soldes d'ouverture (onglet « Soldes d'ouverture »)."""
    balance = compute_balance(conn, only_with_movement=False)
    cr = compute_compte_resultat(conn)

    immo_brutes = sum(b["solde_cloture"] for b in balance if b["classe"] == "2" and int(b["code"]) < 280000)
    amortissements = sum(b["solde_cloture"] for b in balance if b["classe"] == "2" and int(b["code"]) >= 280000)
    immo_nettes = immo_brutes + amortissements

    stocks = stock_initial + _sum_accounts_cloture(balance, COMPTES_STOCK)

    creances = _sum_class(balance, "4", sign="pos")
    treso_actif = _sum_class(balance, "5", sign="pos")
    total_actif = immo_nettes + stocks + creances + treso_actif

    capital = _sum_accounts_cloture(balance, COMPTES_CAPITAL) * -1
    subventions = _sum_accounts_cloture(balance, [COMPTE_SUBVENTIONS]) * -1
    provisions = _sum_accounts_cloture(balance, [COMPTE_PROVISIONS]) * -1
    resultat_net = cr["resultat_net"]
    dettes_financieres = _sum_accounts_cloture(balance, COMPTES_DETTES_FIN) * -1
    dettes_circulantes = -_sum_class(balance, "4", sign="neg")
    treso_passif = -_sum_class(balance, "5", sign="neg")
    total_passif = (capital + subventions + provisions + resultat_net
                     + dettes_financieres + dettes_circulantes + treso_passif)

    return {
        "actif": {
            "Immobilisations brutes": immo_brutes,
            "Amortissements (à déduire)": amortissements,
            "Immobilisations nettes": immo_nettes,
            "Stocks": stocks,
            "Créances et emplois assimilés": creances,
            "Trésorerie actif": treso_actif,
        },
        "total_actif": total_actif,
        "passif": {
            "Capital et réserves": capital,
            "Subventions d'investissement": subventions,
            "Provisions pour risques et charges": provisions,
            "Résultat net de l'exercice": resultat_net,
            "Dettes financières": dettes_financieres,
            "Dettes circulantes": dettes_circulantes,
            "Trésorerie passif": treso_passif,
        },
        "total_passif": total_passif,
        "ecart": total_actif - total_passif,
    }


def compute_grand_livre(conn, compte, tiers=None, date_from=None, date_to=None):
    """Détail chronologique des écritures d'un compte, avec solde cumulé."""
    query = "SELECT * FROM entries WHERE compte = ?"
    params = [compte]
    if tiers:
        query += " AND tiers LIKE ?"
        params.append(f"%{tiers}%")
    if date_from:
        query += " AND date >= ?"
        params.append(date_from)
    if date_to:
        query += " AND date <= ?"
        params.append(date_to)
    query += " ORDER BY date, id"
    rows = [dict(r) for r in conn.execute(query, params).fetchall()]
    solde = 0.0
    for r in rows:
        solde += r["debit"] - r["credit"]
        r["solde_cumule"] = solde
    return rows


# ---------------------------------------------------------------------------
# Stocks
# ---------------------------------------------------------------------------
def compute_stocks(conn):
    balance = compute_balance(conn, only_with_movement=False)
    by_code = {b["code"]: b for b in balance}
    result = []
    for code in COMPTES_STOCK:
        b = by_code.get(code, {"label": get_account_label(conn, code), "debit": 0.0, "credit": 0.0,
                                "solde_ouverture": 0.0})
        initial = get_opening_balance(conn, code)
        entrees, sorties = b["debit"], b["credit"]
        result.append({
            "code": code, "label": b["label"], "stock_initial": initial,
            "entrees": entrees, "sorties": sorties,
            "stock_final": initial + entrees - sorties,
        })
    return result


def set_stock_initial(conn, code, value):
    set_opening_balance(conn, code, value)


# ---------------------------------------------------------------------------
# Production / coûts de fabrication (écritures taguées analytic_code = AN-FAB)
# ---------------------------------------------------------------------------
FLUX_FAB = "AN-FAB"
FAB_POSTES = [
    ("Matières premières et fournitures consommées", ["602000", "604000"]),
    ("Main-d'œuvre directe de production", ["661000", "663000", "664000"]),
    ("Charges indirectes de fabrication", ["624000", "625000", "681000"]),
]


def compute_production(conn):
    balance = compute_balance(conn, only_with_movement=False)

    def net_produit(codes):
        d, c = _sum_accounts(balance, codes)
        return c - d

    ventes = net_produit(["702000", "705000", "706000"])
    stock_d, stock_c = _sum_accounts(balance, ["360000"])
    production_stockee = stock_d - stock_c
    valeur_production = ventes + production_stockee

    postes = []
    total_cout = 0.0
    for label, codes in FAB_POSTES:
        placeholders = ",".join("?" * len(codes))
        row = conn.execute(
            f"SELECT COALESCE(SUM(debit),0) d, COALESCE(SUM(credit),0) c FROM entries "
            f"WHERE compte IN ({placeholders}) AND analytic_code = ?",
            (*codes, FLUX_FAB),
        ).fetchone()
        montant = row["d"] - row["c"]
        postes.append({"label": label, "comptes": ", ".join(codes), "montant": montant})
        total_cout += montant

    return {
        "ventes": ventes,
        "production_stockee": production_stockee,
        "valeur_production": valeur_production,
        "postes_cout": postes,
        "cout_production": total_cout,
        "marge": valeur_production - total_cout,
    }


def compute_tft(conn, treso_ouverture=None):
    """treso_ouverture=None : dérivée automatiquement des soldes d'ouverture des
    comptes de trésorerie (521000/531000/570000/585000). Passez une valeur pour
    la forcer manuellement."""
    balance = compute_balance(conn, only_with_movement=False)
    if treso_ouverture is None:
        treso_ouverture = _sum_accounts_cloture(
            [dict(b, solde_cloture=b["solde_ouverture"]) for b in balance], COMPTES_TRESORERIE)
    treso_debit, treso_credit = _sum_accounts(balance, COMPTES_TRESORERIE)
    variation_totale = treso_debit - treso_credit

    def flux(code):
        d = c = 0.0
        rows = conn.execute(
            "SELECT COALESCE(SUM(debit),0) d, COALESCE(SUM(credit),0) c FROM entries "
            "WHERE compte IN (%s) AND flux_code = ?" % ",".join("?" * len(COMPTES_TRESORERIE)),
            (*COMPTES_TRESORERIE, code),
        ).fetchone()
        return rows["d"] - rows["c"]

    exploitation = flux("FLUX-EXP")
    investissement = flux("FLUX-INV")
    financement = flux("FLUX-FIN")
    non_classes = variation_totale - (exploitation + investissement + financement)

    cloture = treso_ouverture + variation_totale
    return {
        "ouverture": treso_ouverture,
        "exploitation": exploitation,
        "investissement": investissement,
        "financement": financement,
        "non_classes": non_classes,
        "variation": variation_totale,
        "cloture": cloture,
    }


# ---------------------------------------------------------------------------
# Export de la liasse fiscale (.xlsx), mise en page SYSCOHADA système normal
# ---------------------------------------------------------------------------
COMPANY_FIELDS = {
    "societe_nom": "Dénomination sociale",
    "societe_sigle": "Sigle usuel",
    "societe_adresse": "Adresse",
    "societe_ifu": "N° IFU du contribuable",
    "societe_teledeclarant": "N° de télédéclarant (NES)",
    "exercice_clos_le": "Exercice clos le (AAAA-MM-JJ)",
}


def get_company_info(conn):
    return {k: conn.execute("SELECT value FROM settings WHERE key = ?", (k,)).fetchone()
            for k in COMPANY_FIELDS}


def get_company_value(conn, key, default=""):
    row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    return row["value"] if row else default


def set_company_value(conn, key, value):
    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, value))
    conn.commit()


def export_liasse_fiscale(conn, path, stock_initial=0.0, treso_ouverture=0.0):
    """Génère un classeur .xlsx : COUVERTURE, BILAN, RESULTAT, TFT
    (mise en page et codes SYSCOHADA système normal)."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = openpyxl.Workbook()

    bold = Font(bold=True)
    title_font = Font(bold=True, size=13)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    thin = Side(style="thin", color="999999")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    money_fmt = "#,##0"

    def company_row(ws, row=3):
        info = {k: get_company_value(conn, k) for k in COMPANY_FIELDS}
        ws.cell(row=row, column=1, value="Dénomination sociale :")
        ws.cell(row=row, column=3, value=info["societe_nom"])
        ws.cell(row=row + 1, column=1, value="Adresse :")
        ws.cell(row=row + 1, column=3, value=info["societe_adresse"])
        ws.cell(row=row + 2, column=1, value="N° IFU du contribuable :")
        ws.cell(row=row + 2, column=3, value=info["societe_ifu"])
        ws.cell(row=row + 2, column=6, value="Exercice clos le :")
        ws.cell(row=row + 2, column=7, value=info["exercice_clos_le"])
        ws.cell(row=row + 3, column=1, value="N° de télédéclarant (NES) :")
        ws.cell(row=row + 3, column=3, value=info["societe_teledeclarant"])
        for r in range(row, row + 4):
            ws.cell(row=r, column=1).font = bold

    # ---- COUVERTURE ----
    ws = wb.active
    ws.title = "COUVERTURE"
    ws["A1"] = "ÉTATS FINANCIERS — SYSTÈME COMPTABLE OHADA (SYSCOHADA), SYSTÈME NORMAL"
    ws["A1"].font = title_font
    company_row(ws, row=3)
    ws["A9"] = ("Généré automatiquement par l'application Saisie Comptable. Les totaux (AZ, BK, BT, BZ, "
                "CP, DD, DP, DT, DZ) sont calculés directement depuis vos écritures. Le détail par ligne "
                "(AE à AN, CA à CM, DA à DM) est une répartition indicative par plage de comptes — à faire "
                "vérifier par un expert-comptable avant tout dépôt officiel auprès de la DGI.")
    ws["A9"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A9:H9")
    ws.row_dimensions[9].height = 60
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["C"].width = 25

    # ---- BILAN ----
    liasse = compute_liasse_bilan(conn, stock_initial=stock_initial)
    bt = liasse["totaux"]
    ad_net = bt["actif"]["Immobilisations nettes"]
    stocks_net = bt["actif"]["Stocks"]
    creances_net = bt["actif"]["Créances et emplois assimilés"]
    treso_actif_net = bt["actif"]["Trésorerie actif"]
    total_actif = bt["total_actif"]

    ws = wb.create_sheet("BILAN")
    ws["A1"] = "BILAN — SYSTÈME NORMAL"
    ws["A1"].font = title_font
    company_row(ws, row=3)
    headers_row = 8
    ws.cell(row=headers_row, column=1, value="REF").font = bold
    ws.cell(row=headers_row, column=2, value="ACTIF").font = bold
    ws.cell(row=headers_row, column=3, value="BRUT").font = bold
    ws.cell(row=headers_row, column=4, value="AMORT/DEPREC").font = bold
    ws.cell(row=headers_row, column=5, value="NET").font = bold
    ws.cell(row=headers_row, column=7, value="REF").font = bold
    ws.cell(row=headers_row, column=8, value="PASSIF").font = bold
    ws.cell(row=headers_row, column=9, value="NET").font = bold
    for c in range(1, 10):
        ws.cell(row=headers_row, column=c).fill = header_fill

    ad = liasse["actif_detail"]
    actif_lines = [
        ("AE", "Frais de développement et de prospection", ad["AE"]),
        ("AF", "Brevets, licences, logiciels et droits similaires", ad["AF"]),
        ("AG", "Fonds commercial et droit au bail", ad["AG"]),
        ("AH", "Autres immobilisations incorporelles", ad["AH"]),
        ("AJ", "Terrains", ad["AJ"]),
        ("AK", "Bâtiments", ad["AK"]),
        ("AL", "Aménagements, agencements et installations", ad["AL"]),
        ("AM", "Matériel, mobilier et actifs biologiques", ad["AM"]),
        ("AN", "Matériel de transport", ad["AN"]),
        ("AP", "Avances et acomptes versés sur immobilisations", ad["AP"]),
        ("AR", "Titres de participation", ad["AR"]),
        ("AS", "Autres immobilisations financières", ad["AS"]),
    ]
    ac = liasse["actif_circulant_detail"]
    actif_circ_lines = [
        ("BH", "Fournisseurs, avances versées", ac["BH"]),
        ("BI", "Clients", ac["BI"]),
    ]

    pd_ = liasse["passif_detail"]
    passif_lines = [
        ("CA", "Capital", pd_["CA"]),
        ("CD", "Primes liées au capital social", pd_["CD"]),
        ("CF_CG", "Réserves", pd_["CF_CG"]),
        ("CH", "Report à nouveau (+ ou -)", pd_["CH"]),
        ("CJ", "Résultat net de l'exercice", bt["passif"]["Résultat net de l'exercice"]),
        ("CL", "Subventions d'investissement", pd_["CL"]),
        ("CM", "Provisions réglementées", pd_["CM"]),
        ("CP", "TOTAL CAPITAUX PROPRES ET RESSOURCES ASSIMILEES", None),
        ("DA", "Emprunts et dettes financières diverses", pd_["DA"]),
        ("DB", "Dettes de location-acquisition", pd_["DB"]),
        ("DC", "Provisions pour risques et charges", pd_["DC"]),
        ("DD", "TOTAL DETTES FINANCIERES ET RESSOURCES ASSIMILEES", None),
        ("DJ", "Fournisseurs d'exploitation", pd_["DJ"]),
        ("DH", "Clients, avances reçues / Fournisseurs avances (détail)", pd_["DH_avances"]),
        ("DK", "Dettes fiscales et sociales", pd_["DK"]),
        ("DM", "Autres dettes", pd_["DM"]),
    ]

    r = headers_row + 1
    for ref, label, val in actif_lines:
        ws.cell(row=r, column=1, value=ref)
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=5, value=round(val.get("net", val) if isinstance(val, dict) else val))
        ws.cell(row=r, column=5).number_format = money_fmt
        r += 1
    ws.cell(row=r, column=1, value="AZ")
    ws.cell(row=r, column=2, value="TOTAL ACTIF IMMOBILISE").font = bold
    ws.cell(row=r, column=5, value=round(ad_net)).font = bold
    ws.cell(row=r, column=5).number_format = money_fmt
    r += 2
    ws.cell(row=r, column=1, value="BB")
    ws.cell(row=r, column=2, value="STOCKS ET ENCOURS")
    ws.cell(row=r, column=5, value=round(stocks_net)).number_format = money_fmt
    r += 1
    for ref, label, val in actif_circ_lines:
        ws.cell(row=r, column=1, value=ref)
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=5, value=round(val)).number_format = money_fmt
        r += 1
    ws.cell(row=r, column=1, value="BK")
    ws.cell(row=r, column=2, value="TOTAL ACTIF CIRCULANT").font = bold
    ws.cell(row=r, column=5, value=round(stocks_net + creances_net)).font = bold
    ws.cell(row=r, column=5).number_format = money_fmt
    r += 2
    ws.cell(row=r, column=1, value="BT")
    ws.cell(row=r, column=2, value="TOTAL TRESORERIE-ACTIF").font = bold
    ws.cell(row=r, column=5, value=round(treso_actif_net)).font = bold
    ws.cell(row=r, column=5).number_format = money_fmt
    r += 2
    ws.cell(row=r, column=1, value="BZ")
    ws.cell(row=r, column=2, value="TOTAL GENERAL ACTIF").font = bold
    ws.cell(row=r, column=5, value=round(total_actif)).font = bold
    ws.cell(row=r, column=5).number_format = money_fmt
    last_actif_row = r

    r2 = headers_row + 1
    for ref, label, val in passif_lines:
        ws.cell(row=r2, column=7, value=ref)
        ws.cell(row=r2, column=8, value=label)
        if val is not None:
            ws.cell(row=r2, column=9, value=round(val)).number_format = money_fmt
        else:
            ws.cell(row=r2, column=8).font = bold
        r2 += 1
    total_passif = bt["total_passif"]
    ws.cell(row=r2, column=7, value="DZ")
    ws.cell(row=r2, column=8, value="TOTAL GENERAL PASSIF").font = bold
    ws.cell(row=r2, column=9, value=round(total_passif)).font = bold
    ws.cell(row=r2, column=9).number_format = money_fmt
    r2 += 2
    ws.cell(row=r2, column=7, value="Écart Actif - Passif :")
    ws.cell(row=r2, column=9, value=round(total_actif - total_passif)).number_format = money_fmt

    for col, w in zip("ABCDEFGHI", [6, 40, 14, 14, 14, 3, 6, 40, 16]):
        ws.column_dimensions[col].width = w

    # ---- RESULTAT ----
    cr = compute_liasse_resultat(conn)
    ws = wb.create_sheet("RESULTAT")
    ws["A1"] = "COMPTE DE RÉSULTAT — SYSTÈME NORMAL"
    ws["A1"].font = title_font
    company_row(ws, row=3)
    headers_row = 8
    for c, h in zip((1, 2, 5), ("REF", "LIBELLES", "EXERCICE N")):
        ws.cell(row=headers_row, column=c, value=h).font = bold
        ws.cell(row=headers_row, column=c).fill = header_fill

    resultat_lines = [
        ("TA", "Ventes de marchandises", cr["TA"]),
        ("RA", "Achats de marchandises", -cr["RA"]),
        ("XA", "MARGE COMMERCIALE", cr["XA"]),
        ("TB", "Ventes de produits fabriqués", cr["TB"]),
        ("TC", "Travaux, services vendus", cr["TC"]),
        ("TD", "Produits accessoires", cr["TD"]),
        ("XB", "CHIFFRE D'AFFAIRES", cr["XB"]),
        ("TE", "Production stockée (ou déstockage)", cr["TE"]),
        ("TG", "Subventions d'exploitation", cr["TG"]),
        ("TH", "Autres produits", cr["TH"]),
        ("RC", "Achats de matières premières et fournitures liées", -cr["RC"]),
        ("RE", "Autres achats", -cr["RE"]),
        ("RG", "Transports", -cr["RG"]),
        ("RH", "Services extérieurs", -cr["RH"]),
        ("RI", "Impôts et taxes", -cr["RI"]),
        ("RJ", "Autres charges", -cr["RJ"]),
        ("XC", "VALEUR AJOUTEE", cr["XC"]),
        ("RK", "Charges de personnel", -cr["RK"]),
        ("XD", "EXCEDENT BRUT D'EXPLOITATION", cr["XD"]),
        ("RL", "Dotations aux amortissements, provisions et dépréciations", -cr["RL"]),
        ("XE", "RESULTAT D'EXPLOITATION", cr["XE"]),
        ("TK", "Revenus financiers et assimilés", cr["TK"]),
        ("RM", "Frais financiers et charges assimilées", -cr["RM"]),
        ("XF", "RESULTAT FINANCIER", cr["XF"]),
        ("XG", "RESULTAT DES ACTIVITES ORDINAIRES", cr["XG"]),
        ("XH", "RESULTAT HORS ACTIVITES ORDINAIRES (non tracé)", cr["XH"]),
        ("RQ", "Participation des travailleurs (non tracée)", cr["RQ"]),
        ("RS", "Impôts sur le résultat (non tracé — IS à saisir séparément)", cr["RS"]),
        ("XI", "RESULTAT NET", cr["XI"]),
    ]
    bold_refs = {"XA", "XB", "XC", "XD", "XE", "XF", "XG", "XI"}
    r = headers_row + 1
    for ref, label, val in resultat_lines:
        ws.cell(row=r, column=1, value=ref)
        ws.cell(row=r, column=2, value=label)
        cell = ws.cell(row=r, column=5, value=round(val))
        cell.number_format = money_fmt
        if ref in bold_refs:
            ws.cell(row=r, column=2).font = bold
            cell.font = bold
        r += 1
    for col, w in zip("ABCDE", [6, 55, 3, 3, 16]):
        ws.column_dimensions[col].width = w

    # ---- TFT (simplifié, méthode directe) ----
    tft = compute_tft(conn, treso_ouverture=treso_ouverture)
    ws = wb.create_sheet("TFT")
    ws["A1"] = "TABLEAU DES FLUX DE TRÉSORERIE — méthode directe simplifiée"
    ws["A1"].font = title_font
    ws["A2"] = ("Cette version simplifiée (encaissements/décaissements de trésorerie classés EXP/INV/FIN) "
                "ne correspond PAS exactement au format officiel SYSCOHADA (méthode indirecte avec CAFG). "
                "Elle donne une image de la trésorerie mais doit être retravaillée avec un expert-comptable "
                "pour un dépôt officiel.")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 45
    company_row(ws, row=4)
    tft_lines = [
        ("Trésorerie d'ouverture", tft["ouverture"]),
        ("Flux liés aux activités opérationnelles (EXP)", tft["exploitation"]),
        ("Flux liés aux activités d'investissement (INV)", tft["investissement"]),
        ("Flux liés aux activités de financement (FIN)", tft["financement"]),
        ("Flux non classés (à coder)", tft["non_classes"]),
        ("VARIATION NETTE DE TRESORERIE", tft["variation"]),
        ("TRESORERIE DE CLOTURE", tft["cloture"]),
    ]
    r = 10
    for label, val in tft_lines:
        ws.cell(row=r, column=1, value=label)
        cell = ws.cell(row=r, column=3, value=round(val))
        cell.number_format = money_fmt
        r += 1
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["C"].width = 16

    wb.save(path)
    return path


def export_liasse_fiscale_complete(conn, path, stock_initial=0.0):
    """Génère la liasse fiscale COMPLÈTE (mêmes 92 pages, mêmes dimensions que le
    modèle SYSCOHADA système normal fourni) : COUVERTURE/GARDE, BILAN et RESULTAT
    remplis automatiquement depuis vos écritures (soldes de clôture = solde
    d'ouverture + mouvements) ; TFT (officiel, vierge, + un onglet TFT simplifié
    calculé) ; toutes les autres pages (39 notes annexes, ~20 tableaux fiscaux DGI)
    sont conservées avec leur mise en page et leurs dimensions exactes, mais les
    montants qui provenaient du modèle sont effacés (ce ne sont pas vos chiffres)
    pour être complétées manuellement ou par votre expert-comptable."""
    import openpyxl
    from openpyxl.styles import Font

    template_path = os.path.join(_resource_dir(), "etats_financiers_template.xlsx")
    wb = openpyxl.load_workbook(template_path)
    green = Font(color="FF008000")

    # ---- GARDE : identification de l'entité ----
    if "GARDE" in wb.sheetnames:
        g = wb["GARDE"]
        g["D22"] = get_company_value(conn, "societe_nom")
        g["C26"] = get_company_value(conn, "societe_sigle")
        g["C28"] = get_company_value(conn, "societe_adresse")
        g["D30"] = get_company_value(conn, "societe_ifu")
        g["D31"] = get_company_value(conn, "societe_teledeclarant")
        exdate = get_company_value(conn, "exercice_clos_le")
        if exdate:
            g["E17"] = exdate

    # ---- BILAN ----
    liasse = compute_liasse_bilan(conn, stock_initial=stock_initial)
    bt = liasse["totaux"]
    ad = liasse["actif_detail"]
    ac = liasse["actif_circulant_detail"]
    pd_ = liasse["passif_detail"]

    actif_values = {
        "AE": ad["AE"]["net"], "AF": ad["AF"]["net"], "AG": ad["AG"]["net"], "AH": ad["AH"]["net"],
        "AD": ad["AE"]["net"] + ad["AF"]["net"] + ad["AG"]["net"] + ad["AH"]["net"],
        "AJ": ad["AJ"]["net"], "AK": ad["AK"]["net"], "AL": ad["AL"]["net"],
        "AM": ad["AM"]["net"], "AN": ad["AN"]["net"],
        "AI": ad["AJ"]["net"] + ad["AK"]["net"] + ad["AL"]["net"] + ad["AM"]["net"] + ad["AN"]["net"],
        "AP": ad["AP"]["net"], "AR": ad["AR"]["net"], "AS": ad["AS"]["net"],
        "AZ": bt["actif"]["Immobilisations nettes"],
        "BB": bt["actif"]["Stocks"],
        "BH": ac["BH"], "BI": ac["BI"],
        "BK": bt["actif"]["Stocks"] + bt["actif"]["Créances et emplois assimilés"],
        "BT": bt["actif"]["Trésorerie actif"],
        "BZ": bt["total_actif"],
    }
    passif_values = {
        "CA": pd_["CA"], "CD": pd_["CD"], "CF": pd_["CF_CG"], "CH": pd_["CH"],
        "CJ": bt["passif"]["Résultat net de l'exercice"],
        "CL": pd_["CL"], "CM": pd_["CM"],
        "CP": (pd_["CA"] + pd_["CD"] + pd_["CF_CG"] + pd_["CH"]
               + bt["passif"]["Résultat net de l'exercice"] + pd_["CL"] + pd_["CM"]),
        "DA": pd_["DA"], "DB": pd_["DB"], "DC": pd_["DC"],
        "DD": pd_["DA"] + pd_["DB"] + pd_["DC"],
        "DJ": pd_["DJ"], "DH": pd_["DH_avances"], "DK": pd_["DK"], "DM": pd_["DM"],
        "DP": pd_["DJ"] + pd_["DH_avances"] + pd_["DK"] + pd_["DM"],
        "DT": bt["passif"]["Trésorerie passif"],
        "DZ": bt["total_passif"],
    }

    if "BILAN" in wb.sheetnames:
        ws = wb["BILAN"]
        ws["C3"] = get_company_value(conn, "societe_nom")
        ws["C4"] = get_company_value(conn, "societe_adresse")
        ws["C5"] = get_company_value(conn, "societe_ifu")
        for ref, row in {
            "AD": 11, "AE": 12, "AF": 13, "AG": 14, "AH": 15, "AI": 16, "AJ": 17, "AK": 18,
            "AL": 19, "AM": 20, "AN": 21, "AP": 22, "AQ": 23, "AR": 24, "AS": 25, "AZ": 26,
            "BB": 28, "BH": 30, "BI": 31, "BK": 33, "BT": 37, "BZ": 39,
        }.items():
            if ref in actif_values:
                cell = ws.cell(row=row, column=8, value=round(actif_values[ref]))
                cell.font = green
        for ref, row in {
            "CA": 11, "CD": 13, "CF": 15, "CH": 17, "CJ": 18, "CL": 19, "CM": 20, "CP": 21,
            "DA": 22, "DB": 23, "DC": 24, "DD": 25, "DH": 27, "DJ": 29, "DK": 30, "DM": 31,
            "DP": 33, "DT": 36, "DZ": 39,
        }.items():
            if ref in passif_values:
                cell = ws.cell(row=row, column=13, value=round(passif_values[ref]))
                cell.font = green
        ws.cell(row=40, column=13, value="=H39-M39")  # écart de contrôle

    # ---- RESULTAT ----
    cr = compute_liasse_resultat(conn)
    if "RESULTAT" in wb.sheetnames:
        ws = wb["RESULTAT"]
        row_map = {"TA": 11, "RA": 12, "XA": 14, "TB": 15, "TC": 16, "TD": 17, "XB": 18,
                   "TE": 19, "TG": 21, "TH": 22, "RC": 24, "RE": 26, "RG": 28, "RH": 29,
                   "RI": 30, "RJ": 31, "XC": 32, "RK": 33, "XD": 34, "RL": 36, "XE": 37,
                   "TK": 38, "RM": 41, "XF": 43, "XG": 44, "XH": 49, "RQ": 50, "RS": 51, "XI": 52}
        sign_negative = {"RA", "RC", "RE", "RG", "RH", "RI", "RJ", "RK", "RL", "RM", "RQ", "RS"}
        for ref, row in row_map.items():
            val = cr.get(ref, 0.0)
            if ref in sign_negative:
                val = -abs(val)
            cell = ws.cell(row=row, column=9, value=round(val))
            cell.font = green

    # ---- TFT : ajoute un onglet supplémentaire avec notre calcul simplifié ----
    tft = compute_tft(conn)
    if "TFT" in wb.sheetnames:
        ws = wb["TFT"]
        ws["I10"] = round(tft["ouverture"])
        ws["I10"].font = green
        ws["A44"] = ("Feuille officielle laissée vierge (méthode indirecte avec CAFG non calculée "
                     "automatiquement) — voir l'onglet « TFT (simplifie) » pour un calcul indicatif.")
    ws_simple = wb.create_sheet("TFT (simplifie)")
    ws_simple["A1"] = "TFT SIMPLIFIÉ (méthode directe, calculé automatiquement)"
    ws_simple["A1"].font = Font(bold=True, size=12)
    lines = [
        ("Trésorerie d'ouverture", tft["ouverture"]),
        ("Flux liés aux activités opérationnelles (EXP)", tft["exploitation"]),
        ("Flux liés aux activités d'investissement (INV)", tft["investissement"]),
        ("Flux liés aux activités de financement (FIN)", tft["financement"]),
        ("Flux non classés (à coder)", tft["non_classes"]),
        ("VARIATION NETTE DE TRESORERIE", tft["variation"]),
        ("TRESORERIE DE CLOTURE", tft["cloture"]),
    ]
    for i, (label, val) in enumerate(lines):
        ws_simple.cell(row=3 + i, column=1, value=label)
        ws_simple.cell(row=3 + i, column=3, value=round(val))
    ws_simple.column_dimensions["A"].width = 45

    # ---- Toutes les autres pages : structure/dimensions conservées, valeurs
    #      chiffrées (issues du modèle GCM) effacées pour éviter toute confusion ----
    skip = {"GARDE", "BILAN", "RESULTAT", "TFT", "TFT (simplifie)"}
    for name in wb.sheetnames:
        if name in skip:
            continue
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.value = None

    wb.save(path)
    return path


if __name__ == "__main__":
    # Petit auto-test en ligne de commande (sans Tkinter).
    conn = get_connection(":memory:" if False else "test_core.db")
    print("Comptes chargés :", conn.execute("SELECT COUNT(*) FROM accounts").fetchone()[0])

    add_entry(conn, str(date.today()), "FA-0001", "AC", "601000", "", "Achat marchandises", 1000, 0)
    add_entry(conn, str(date.today()), "FA-0001", "AC", "445200", "", "TVA récupérable", 200, 0)
    add_entry(conn, str(date.today()), "FA-0001", "AC", "401000", "Ets Dupont", "Facture FA-0001", 0, 1200)
    add_entry(conn, str(date.today()), "FV-0001", "VE", "411000", "Société ABC", "Facture FV-0001", 1180, 0)
    add_entry(conn, str(date.today()), "FV-0001", "VE", "701000", "", "Vente marchandises", 0, 1180)

    d, c = totals_debit_credit(conn)
    print("Total débit / crédit :", d, c, "Équilibré :", d == c)

    print("\n--- Balance ---")
    for b in compute_balance(conn):
        print(b)

    print("\n--- Compte de résultat ---")
    cr = compute_compte_resultat(conn)
    print("Résultat net :", cr["resultat_net"])

    print("\n--- Bilan ---")
    bilan = compute_bilan(conn)
    print("Total actif :", bilan["total_actif"], "Total passif :", bilan["total_passif"], "Écart :", bilan["ecart"])

    print("\n--- TFT ---")
    print(compute_tft(conn))

    print("\n--- Grand livre (411000) ---")
    for r in compute_grand_livre(conn, "411000"):
        print(r)

    print("\n--- Stocks ---")
    for s in compute_stocks(conn):
        print(s)

    print("\n--- Production ---")
    print(compute_production(conn))

    conn.close()
    os.remove("test_core.db")
